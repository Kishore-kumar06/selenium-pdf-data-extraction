from dotenv import load_dotenv
import os
import datetime
from openpyxl import Workbook, load_workbook
from utils.logfiles import setup_logger
import time
import shutil

logger = setup_logger("file_and_tracker")
load_dotenv()

class File_And_Tracker:
    def __init__(self, main_path):
        self.main_path = os.path.abspath(main_path)
        # self.pipelines = pipelines

    # This function creates an Excel file to track the status of downloaded files. It takes the main path, sub-path, pipeline name, company name, tariff program, effective status, file status, and time taken as input. If the tracker file for the current date already exists, it appends a new entry to it; otherwise, it creates a new tracker file with headers and the first entry.
    def create_excel_tracker_files(self, pipeline, company_name, tariff_program, is_effective, file_status, time_taken):
        try:
            tracker_dir = os.getenv("TRACEBACK_FILE")
            tracker_path = os.path.abspath(os.path.join(self.main_path, tracker_dir))
        
            # Ensure directory exists
            os.makedirs(tracker_path, exist_ok=True)

            today = datetime.date.today().strftime("%Y-%m-%d")
            tracker_file = os.path.join(tracker_path, f"tracker_file_{today}.xlsx")

            new_entry = [pipeline, company_name, tariff_program, is_effective, file_status, round(time_taken,2)]

            max_attempts = 3
            for attempt in range(1, max_attempts + 1):
                try:
                    if os.path.exists(tracker_file):
                        wb = load_workbook(tracker_file)
                        ws = wb.active
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.append(['Pipeline', 'Company Name', 'Tariff Title', 'Effective Status', 'File Status', 'Time Taken'])
                    
                    ws.append(new_entry)
                    wb.save(tracker_file)
                    logger.info(f"Tracker updated successfully: {tracker_file}.")
                    break

                except PermissionError as e:
                    logger.error(f"Error recording tracker file: {e}.")
                    if attempt == max_attempts:
                        raise
                    time.sleep(2)  

        except Exception as e:
            logger.error(f"Error creating tracker file: {e}.")

    
    # This function creates a folder for each pipeline inside the specified main path and sub-path to store the downloaded pdf files.
    def create_pipeline_folder(self, pipeline):
        try:
            download_directory = os.getenv("DOWNLOAD_DIR")
            output_path = os.path.abspath(os.path.join(self.main_path, download_directory))

            pipeline_folder = os.path.join(output_path, pipeline)

            if not os.path.exists(pipeline_folder):
                os.makedirs(pipeline_folder, exist_ok=True)
            else:
                logger.error(f"Pipeline folder {pipeline_folder} already exist.")
            
            return pipeline_folder.replace("\\", "/")  # Return the path with forward slashes for consistency
            
        except Exception as e:
            logger.error(f"Error occured while creating pipeline folder: {e}.")
            raise

   
    def get_latest_file(self, source_file_path, destination_file_path, pipeline, timeout=30):
        try:
            target_path = os.path.abspath(source_file_path)
            if not os.path.exists(target_path):
                raise FileNotFoundError(f"Folder not found: {target_path}")

            start_time = time.time()
            logger.info(f"Polling download validation directory for completion: {target_path}")

            while time.time() - start_time < timeout:

                all_items = os.listdir(target_path)
                has_active_stream = any(item.endswith(".crdownload") or item.endswith(".tmp") for item in all_items)
            
                valid_files = [os.path.join(target_path, f) for f in all_items if os.path.isfile(os.path.join(target_path, f)) and not (f.endswith(".crdownload") or f.endswith(".crdownload"))]

                if valid_files and not has_active_stream:
                    break

                time.sleep(1)

            else:
                raise TimeoutError(f"File for {pipeline} pipeline has not downloaded within the time")
            
            files = [os.path.join(target_path, f) for f in os.listdir(target_path) if os.path.isfile(os.path.join(target_path, f))]

            if not files:
                raise ValueError("No files available in directory {file_path}.")

            # Get latest file
            downloaded_file = max(files, key=os.path.getctime)
            # Extract extension
            _, file_extension = os.path.splitext(downloaded_file)

            # New file name
            new_file_name = f"{pipeline}{file_extension}"
            # New full path
            latest_file_path = os.path.join(target_path, new_file_name)

            if os.path.exists(latest_file_path):
                os.remove(latest_file_path)

            # Rename file
            os.rename(downloaded_file, latest_file_path)

            shutil.move(latest_file_path, destination_file_path)

            return destination_file_path
        except Exception as e:
            logger.exception(f"Error getting latest file: {e}.")
            return None
        
    def get_processed_files(self):
        try:
            processed_file_directory = os.getenv("PROCESSED_FILE_PATH")

            if not processed_file_directory:
                os.makedirs(processed_file_directory, exist_ok=True)

            today = datetime.date.today().strftime("%Y-%m-%d")
            processed_file = f"processed_file_{today}.txt"

            tracker_file_path = os.path.join(processed_file_directory, processed_file)

            if not os.path.exists(tracker_file_path):
                return set()
            
            with open(tracker_file_path, 'r') as f:
                if f.readline() == "":
                    raise ValueError(f"No pipelenes have been downloaded.")
                
                return set(line.strip() for line in f if line.strip())
        except Exception as e:
            logger.exception(f"Processed pipelines file do not exist at {processed_file_directory} directory. {e}")
            raise FileNotFoundError(f"Processed pipelines file do not exist at {processed_file_directory} directory.")  


    def write_downloaded_pipelines(self, pipeline):
        try:
            processed_file_directory = os.getenv("PROCESSED_FILE_PATH")

            if not processed_file_directory:
                os.makedirs(processed_file_directory, exist_ok=True)

            today = datetime.date.today().strftime("%Y-%m-%d")
            processed_file = f"processed_file_{today}.txt"

            tracker_file_path = os.path.join(processed_file_directory, processed_file)

            with open(tracker_file_path, 'a') as f:
                f.write(f"{pipeline}\n")

        except Exception as e:
            logger.exception(f"Processed pipelines file do not exist at {processed_file_directory} directory. {e}")
            raise FileNotFoundError(f"Processed pipelines file do not exist at {processed_file_directory} directory.")  

